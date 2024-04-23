import openpyxl
from openpyxl.styles import PatternFill

def compare_and_highlight(file1, file2):
    # Load the Excel files
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Get common sheet names
    common_sheet_names = set(wb1.sheetnames).intersection(wb2.sheetnames)

    for sheet_name in common_sheet_names:
        # Get sheets from both workbooks
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # Iterate through each cell and compare values
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell1 = sheet1.cell(row=row, column=col)
                cell2 = sheet2.cell(row=row, column=col)

                # Compare cell values
                if cell1.value != cell2.value:
                    # If values are different, set background color to red in the first workbook
                    fill_red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    cell1.fill = fill_red

                    # Set background color to green in the second workbook
                    fill_green = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                    cell2.fill = fill_green

    # Save the modified workbooks
    wb1.save(file1)
    wb2.save(file2)

    print(f'Differences highlighted in {file1} and {file2}')

# Example usage:
file1 = './uploads/1.xlsx'
file2 = './uploads/2.xlsx'
compare_and_highlight(file1, file2)
